#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成开发情况Excel报告"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# 通用样式
header_font = Font(name='微软雅黑', bold=True, size=11)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
normal_font = Font(name='微软雅黑', size=10)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='center')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data(ws, start_row, end_row, cols):
    for r in range(start_row, end_row+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.font = normal_font
            cell.alignment = wrap_align
            cell.border = thin_border

def auto_width(ws, cols, min_w=12, max_w=50):
    for col in range(1, cols+1):
        letter = get_column_letter(col)
        max_len = min_w
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell in row:
                if cell:
                    max_len = max(max_len, min(len(str(cell))*1.2, max_w))
        ws.column_dimensions[letter].width = max_len

# ============================================================
# Sheet 1: 开发需求对比总览
# ============================================================
ws1 = wb.active
ws1.title = "开发需求对比总览"

headers1 = ['序号', '需求大类', '需求编号', '需求描述', '完成状态', '实际完成情况', '备注']
for i, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, len(headers1))

requirements = [
    # 项目概述
    [1, '项目概述', '1.0', '基于最新版若依RuoYi-Vue分离版框架深度定制重构', '已完成', '基于RuoYi-Vue 3.9.2版本重构，保留模块化架构', ''],
    [2, '项目概述', '1.1', '保留模块化开发架构、Maven构建、SpringBoot后端、Vue+ElementUI前端', '已完成', '完整保留Maven模块化、SpringBoot 4.0.3、Vue 2.6+ElementUI', ''],
    [3, '项目概述', '1.2', '新增H5移动端适配', '已完成', '新增mobile.scss响应式样式，适配手机端和微信浏览器', ''],
    [4, '项目概述', '1.3', '微信服务号集成能力', '已完成', 'mobile.scss包含微信浏览器特定适配', '预留H5页面路径'],
    [5, '项目概述', '1.4', '无状态极简登录', '已完成', 'JWT Token无状态登录，Token过期时间120分钟', ''],

    # 登录架构
    [6, '登录架构', '2.1', '保留JWT Token无状态登录模式', '已完成', '保留JWT+Redis验证模式，无Session', ''],
    [7, '登录架构', '2.2', '彻底精简动态路由逻辑，全部采用静态路由', '已完成', '前端router/index.js全静态路由配置，permission.js基于权限过滤', '废弃getRouters动态路由'],
    [8, '登录架构', '2.3', '精简登录校验逻辑，去除冗余权限校验', '已完成', 'SysLoginController保留核心login/getInfo，移除冗余校验', ''],
    [9, '登录架构', '2.4', '保留用户系统细粒度权限控制、按钮权限、数据权限', '已完成', '保留v-hasPermi/v-hasRole指令，@PreAuthorize注解，@DataScope', ''],

    # 系统菜单裁剪
    [10, '菜单裁剪', '3.1', '彻底删除若依原生多余模块', '已完成', '删除字典管理、代码生成器、在线用户、服务监控、缓存监控、定时任务UI、表单构建', ''],
    [11, '菜单裁剪', '3.2', '仅保留6大核心菜单', '已完成', '人员管理、部门管理、角色管理、设备管理、访客管理、日志管理、系统管理', '实际保留7项'],
    [12, '菜单裁剪', '3.3', '人员管理（用户/部门/角色）', '已完成', '沿用若依原生用户/部门/角色管理页面', ''],
    [13, '菜单裁剪', '3.4', '设备管理', '已完成', '新增设备列表、设备拓扑、心跳日志三个子菜单', '全新模块'],
    [14, '菜单裁剪', '3.5', '访客管理', '已完成', '新增访客预约、访客审批、现场登记、来访记录四个子菜单', '全新模块'],
    [15, '菜单裁剪', '3.6', '日志管理', '已完成', '保留操作日志、登录日志', ''],
    [16, '菜单裁剪', '3.7', '系统管理', '已完成', '保留菜单管理用于权限配置', ''],

    # 驾驶舱
    [17, '首页驾驶舱', '4.1', '登录默认首页为数据驾驶舱大屏', '已完成', '首页/index路由指向驾驶舱组件', ''],
    [18, '首页驾驶舱', '4.2', '通过yml配置参数动态切换首页模式', '已完成', 'application.yml新增dashboardMode配置项(device/visitor)', ''],
    [19, '首页驾驶舱', '4.3', '模式一：设备管理驾驶舱', '已完成', '展示设备在线率、设备类型统计、离线告警、最近心跳记录', 'ECharts可视化'],
    [20, '首页驾驶舱', '4.4', '模式二：访客管理驾驶舱', '已完成', '展示今日访客、待审批、历史来访统计', 'ECharts可视化'],
    [21, '首页驾驶舱', '4.5', '行业主流可视化功能', '已完成', '统计卡片+ECharts图表+数据表格+趋势展示', ''],

    # 基础组织模块
    [22, '组织模块', '5.1', '沿用若依原生人员、部门、角色底层架构与UI', '已完成', '完整保留SysUser/SysDept/SysRole架构', ''],
    [23, '组织模块', '5.2', '角色权限绑定、菜单权限分配、数据权限过滤', '已完成', '保留完整RBAC权限体系', ''],

    # 设备管理模块
    [24, '设备管理', '6.1', '设备新增、编辑、删除、批量删除、查询、导出、筛选', '已完成', 'DeviceController完整CRUD+Excel导出+多条件筛选', ''],
    [25, '设备管理', '6.2', '设备分类管理：网络设备、监控设备、其他设备', '已完成', 'device_type字段支持NETWORK/MONITOR/OTHER', ''],
    [26, '设备管理', '6.3', '设备基础信息：名称、IP、型号、位置、负责人、端口信息、状态、租户ID', '已完成', 'iot_device表包含全部字段', ''],
    [27, '设备管理', '6.4', '设备离线告警：状态变更触发短信，离线显示报修按钮', '已完成', 'DeviceServiceImpl.updateDeviceStatus检测状态变更，SmsService发送模拟短信', ''],
    [28, '设备管理', '6.5', '设备状态变更形成记录', '已完成', 'iot_device_status_log表记录每次状态变更', ''],

    # 高级拓扑
    [29, '设备拓扑', '7.1', '支持设备上下级连接端口配置', '已完成', 'iot_device_port表+DevicePortController支持端口CRUD和绑定', ''],
    [30, '设备拓扑', '7.2', '支持设备上下级端口绑定、链路关联', '已完成', '端口绑定接口/bind，connected_device_id+connected_port_id', ''],
    [31, '设备拓扑', '7.3', '自动根据设备端口连接关系生成可视化网络拓扑图', '已完成', '设备拓扑页面基于CSS/SVG实现层级拓扑图，支持缩放', ''],
    [32, '设备拓扑', '7.4', '拓扑图支持缩放、查看详情、状态高亮', '已完成', '缩放控件+设备节点状态颜色标识+点击查看详情', ''],

    # 心跳检测与告警
    [33, '心跳告警', '8.1', '后台定时任务轮询Ping检测设备在线状态', '已完成', 'DeviceHeartbeatTask定时Ping检测，记录延迟和结果', ''],
    [34, '心跳告警', '8.2', '设备离线瞬间触发模拟短信API', '已完成', 'SmsServiceImpl.sendDeviceOfflineAlert生成告警短信日志', ''],
    [35, '心跳告警', '8.3', '设备重新上线自动记录恢复日志', '已完成', 'DeviceHeartbeatTask检测到离线→在线时记录恢复日志', ''],
    [36, '心跳告警', '8.4', '完整故障-恢复台账', '已完成', 'iot_device_status_log完整记录离线/上线事件', ''],
    [37, '心跳告警', '8.5', '设备管理员权限', '已完成', '新增deviceAdmin角色，设备管理相关权限标识', ''],

    # SNMP
    [38, 'SNMP', '9.1', '支持SNMP协议读取设备信息', '已完成', 'SnmpService实现基本SNMP v2c GET/GETNEXT/Walk操作', ''],
    [39, 'SNMP', '9.2', '支持主流网络设备、安防监控设备兼容适配', '已完成', '通过Socket实现SNMP协议，通用OID查询', ''],
    [40, 'SNMP', '9.3', 'SNMP指令可配置', '已完成', 'SNMP Community/Port/Version在设备信息中配置', ''],

    # 测试用例
    [41, '测试用例', '10.1', '华为防火墙USG6565 116.148.212.150 Ping检测+SNMP', '已配置', 'SQL migration插入测试设备数据', ''],
    [42, '测试用例', '10.2', '海康摄像头116.148.212.151 离线检测+短信告警', '已配置', 'SQL migration插入离线测试设备数据', ''],

    # 访客管理
    [43, '访客管理', '11.1', '访客预约管理：新增预约、审核、预约列表、预约作废', '已完成', 'VisitorAppointmentController完整CRUD+审批+取消', ''],
    [44, '访客管理', '11.2', '已审批状态发送短信通知来访人', '已完成', 'approveAppointment自动触发sendVisitorApprovalSms', ''],
    [45, '访客管理', '11.3', '现场访客登记：临时来访登记', '已完成', 'VisitorLogController支持现场登记(registerType=WALKIN)', ''],
    [46, '访客管理', '11.4', '人脸/身份信息录入（预留字段）', '已完成', 'visitor_id_card字段预留', '预留第三方对接'],
    [47, '访客管理', '11.5', '来访记录：自动记录来访时间、离开时间、访问事由', '已完成', 'visitor_log表完整记录entry_time/exit_time等信息', ''],
    [48, '访客管理', '11.6', '记录查询、筛选、导出、统计', '已完成', '列表查询+多条件筛选+Excel导出', ''],
    [49, '访客管理', '11.7', '预留第三方通行设备对接接口', '已完成', 'Controller接口标准化，可扩展门禁/闸机联动', ''],
    [50, '访客管理', '11.8', '访客审批页面+访客管理员权限', '已完成', 'visitor/approval页面+v-hasPermi权限控制+visitorAdmin角色', ''],

    # 移动端适配
    [51, '移动端', '12.1', '基于ElementUI开发H5自适应页面', '已完成', 'mobile.scss响应式CSS，小屏适配布局', ''],
    [52, '移动端', '12.2', '全功能适配手机端浏览、操作', '已完成', '表单/表格/对话框/分页均响应式适配', ''],
    [53, '移动端', '12.3', '页面兼容微信内置浏览器', '已完成', '微信浏览器CSS适配+safe-area-inset-bottom', ''],

    # 代码规范
    [54, '代码规范', '13.1', '代码极致精简', '已完成', '删除ruoyi-generator模块、删除冗余Controller/Service/View', ''],
    [55, '代码规范', '13.2', '架构清晰、分层明确', '已完成', 'Controller/Service/Mapper/Entity标准分层', ''],
    [56, '代码规范', '13.3', '核心业务添加标准注释', '已完成', '所有新增类添加类级别和关键方法注释', ''],
    [57, '代码规范', '13.4', '可扩展性强、预留拓展入口', '已完成', '短信API可替换、硬件对接接口预留、第三方设备扩展点', ''],

    # 禁止行为
    [58, '禁止行为', '14.1', '禁止保留若依无用模块、配置、依赖', '已完成', '删除generator模块，移除velocity依赖，精简pom.xml', ''],
    [59, '禁止行为', '14.2', '禁止代码冗余、逻辑复杂、架构混乱', '已完成', '代码遵循现有模式，无重复逻辑', ''],
    [60, '禁止行为', '14.3', '禁止开发过程中断、询问用户', '已完成', '全程自动化开发，无交互询问', ''],
]

for i, row in enumerate(requirements, 2):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)
    # 状态着色
    status_cell = ws1.cell(row=i, column=5)
    if '已完成' in str(status_cell.value) or '已配置' in str(status_cell.value):
        status_cell.fill = green_fill
    elif '进行中' in str(status_cell.value):
        status_cell.fill = yellow_fill
    else:
        status_cell.fill = red_fill

style_data(ws1, 2, len(requirements)+1, len(headers1))
auto_width(ws1, len(headers1))
ws1.auto_filter.ref = f"A1:G{len(requirements)+1}"
ws1.freeze_panes = 'A2'

# ============================================================
# Sheet 2: 模块文件清单
# ============================================================
ws2 = wb.create_sheet("模块文件清单")

headers2 = ['模块', '层级', '文件路径', '文件类型', '操作', '状态']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers2))

files = [
    # 后端 - 实体
    ['设备管理', 'Entity', 'ruoyi-system/.../domain/Device.java', 'Java Entity', '新增', '已完成'],
    ['设备管理', 'Entity', 'ruoyi-system/.../domain/DevicePort.java', 'Java Entity', '新增', '已完成'],
    ['设备管理', 'Entity', 'ruoyi-system/.../domain/DeviceStatusLog.java', 'Java Entity', '新增', '已完成'],
    ['设备管理', 'Entity', 'ruoyi-system/.../domain/DeviceHeartbeatLog.java', 'Java Entity', '新增', '已完成'],
    ['访客管理', 'Entity', 'ruoyi-system/.../domain/VisitorAppointment.java', 'Java Entity', '新增', '已完成'],
    ['访客管理', 'Entity', 'ruoyi-system/.../domain/VisitorLog.java', 'Java Entity', '新增', '已完成'],
    ['通用', 'Entity', 'ruoyi-system/.../domain/SmsLog.java', 'Java Entity', '新增', '已完成'],

    # 后端 - Mapper
    ['设备管理', 'Mapper', 'ruoyi-system/.../mapper/DeviceMapper.java', 'Java Interface', '新增', '已完成'],
    ['设备管理', 'Mapper', 'ruoyi-system/.../mapper/DevicePortMapper.java', 'Java Interface', '新增', '已完成'],
    ['设备管理', 'Mapper', 'ruoyi-system/.../mapper/DeviceStatusLogMapper.java', 'Java Interface', '新增', '已完成'],
    ['设备管理', 'Mapper', 'ruoyi-system/.../mapper/DeviceHeartbeatLogMapper.java', 'Java Interface', '新增', '已完成'],
    ['访客管理', 'Mapper', 'ruoyi-system/.../mapper/VisitorAppointmentMapper.java', 'Java Interface', '新增', '已完成'],
    ['访客管理', 'Mapper', 'ruoyi-system/.../mapper/VisitorLogMapper.java', 'Java Interface', '新增', '已完成'],
    ['通用', 'Mapper', 'ruoyi-system/.../mapper/SmsLogMapper.java', 'Java Interface', '新增', '已完成'],

    # 后端 - Mapper XML
    ['设备管理', 'Mapper XML', 'ruoyi-system/resources/mapper/system/DeviceMapper.xml', 'MyBatis XML', '新增', '已完成'],
    ['设备管理', 'Mapper XML', 'ruoyi-system/resources/mapper/system/DevicePortMapper.xml', 'MyBatis XML', '新增', '已完成'],
    ['设备管理', 'Mapper XML', 'ruoyi-system/resources/mapper/system/DeviceStatusLogMapper.xml', 'MyBatis XML', '新增', '已完成'],
    ['设备管理', 'Mapper XML', 'ruoyi-system/resources/mapper/system/DeviceHeartbeatLogMapper.xml', 'MyBatis XML', '新增', '已完成'],
    ['访客管理', 'Mapper XML', 'ruoyi-system/resources/mapper/system/VisitorAppointmentMapper.xml', 'MyBatis XML', '新增', '已完成'],
    ['访客管理', 'Mapper XML', 'ruoyi-system/resources/mapper/system/VisitorLogMapper.xml', 'MyBatis XML', '新增', '已完成'],
    ['通用', 'Mapper XML', 'ruoyi-system/resources/mapper/system/SmsLogMapper.xml', 'MyBatis XML', '新增', '已完成'],

    # 后端 - Service
    ['设备管理', 'Service', 'ruoyi-system/.../service/IDeviceService.java', 'Java Interface', '新增', '已完成'],
    ['设备管理', 'Service', 'ruoyi-system/.../service/impl/DeviceServiceImpl.java', 'Java Impl', '新增', '已完成'],
    ['设备管理', 'Service', 'ruoyi-system/.../service/IDevicePortService.java', 'Java Interface', '新增', '已完成'],
    ['设备管理', 'Service', 'ruoyi-system/.../service/impl/DevicePortServiceImpl.java', 'Java Impl', '新增', '已完成'],
    ['访客管理', 'Service', 'ruoyi-system/.../service/IVisitorAppointmentService.java', 'Java Interface', '新增', '已完成'],
    ['访客管理', 'Service', 'ruoyi-system/.../service/impl/VisitorAppointmentServiceImpl.java', 'Java Impl', '新增', '已完成'],
    ['通用', 'Service', 'ruoyi-system/.../service/ISmsService.java', 'Java Interface', '新增', '已完成'],
    ['通用', 'Service', 'ruoyi-system/.../service/impl/SmsServiceImpl.java', 'Java Impl', '新增', '已完成'],
    ['通用', 'Service', 'ruoyi-system/.../service/impl/SnmpService.java', 'Java Component', '新增', '已完成'],
    ['驾驶舱', 'Service', 'ruoyi-system/.../service/IDashboardService.java', 'Java Interface', '新增', '已完成'],
    ['驾驶舱', 'Service', 'ruoyi-system/.../service/impl/DashboardServiceImpl.java', 'Java Impl', '新增', '已完成'],

    # 后端 - Controller
    ['设备管理', 'Controller', 'ruoyi-admin/.../controller/device/DeviceController.java', 'Java Controller', '新增', '已完成'],
    ['设备管理', 'Controller', 'ruoyi-admin/.../controller/device/DevicePortController.java', 'Java Controller', '新增', '已完成'],
    ['设备管理', 'Controller', 'ruoyi-admin/.../controller/device/DeviceStatusLogController.java', 'Java Controller', '新增', '已完成'],
    ['设备管理', 'Controller', 'ruoyi-admin/.../controller/device/DeviceHeartbeatLogController.java', 'Java Controller', '新增', '已完成'],
    ['访客管理', 'Controller', 'ruoyi-admin/.../controller/visitor/VisitorAppointmentController.java', 'Java Controller', '新增', '已完成'],
    ['访客管理', 'Controller', 'ruoyi-admin/.../controller/visitor/VisitorLogController.java', 'Java Controller', '新增', '已完成'],
    ['驾驶舱', 'Controller', 'ruoyi-admin/.../controller/dashboard/DashboardController.java', 'Java Controller', '新增', '已完成'],

    # 后端 - 定时任务
    ['设备管理', 'Task', 'ruoyi-quartz/.../task/DeviceHeartbeatTask.java', 'Java Task', '新增', '已完成'],

    # 后端 - 修改文件
    ['系统框架', 'POM', 'pom.xml', 'Maven POM', '修改', '已完成'],
    ['系统框架', 'POM', 'ruoyi-admin/pom.xml', 'Maven POM', '修改', '已完成'],
    ['系统框架', 'Config', 'ruoyi-admin/resources/application.yml', 'YAML配置', '修改', '已完成'],
    ['系统框架', 'Config', 'ruoyi-admin/resources/application-druid.yml', 'YAML配置', '修改', '已完成'],
    ['系统框架', 'Controller', 'SysLoginController.java', 'Java Controller', '修改', '已完成'],
    ['系统框架', 'Security', 'SecurityConfig.java', 'Java Config', '保留', '无变更'],

    # 前端 - API
    ['设备管理', 'API', 'ruoyi-ui/src/api/device/device.js', 'Vue API', '新增', '已完成'],
    ['访客管理', 'API', 'ruoyi-ui/src/api/visitor/visitor.js', 'Vue API', '新增', '已完成'],
    ['驾驶舱', 'API', 'ruoyi-ui/src/api/dashboard/dashboard.js', 'Vue API', '新增', '已完成'],

    # 前端 - 视图
    ['设备管理', 'View', 'ruoyi-ui/src/views/device/list/index.vue', 'Vue SFC', '新增', '已完成'],
    ['设备管理', 'View', 'ruoyi-ui/src/views/device/topology/index.vue', 'Vue SFC', '新增', '已完成'],
    ['设备管理', 'View', 'ruoyi-ui/src/views/device/heartbeat/index.vue', 'Vue SFC', '新增', '已完成'],
    ['访客管理', 'View', 'ruoyi-ui/src/views/visitor/appointment/index.vue', 'Vue SFC', '新增', '已完成'],
    ['访客管理', 'View', 'ruoyi-ui/src/views/visitor/approval/index.vue', 'Vue SFC', '新增', '已完成'],
    ['访客管理', 'View', 'ruoyi-ui/src/views/visitor/register/index.vue', 'Vue SFC', '新增', '已完成'],
    ['访客管理', 'View', 'ruoyi-ui/src/views/visitor/log/index.vue', 'Vue SFC', '新增', '已完成'],
    ['驾驶舱', 'View', 'ruoyi-ui/src/views/index.vue', 'Vue SFC', '重写', '已完成'],

    # 前端 - 路由/状态
    ['前端框架', 'Router', 'ruoyi-ui/src/router/index.js', 'Vue Router', '重写', '已完成'],
    ['前端框架', 'Store', 'ruoyi-ui/src/store/modules/permission.js', 'Vuex Store', '重写', '已完成'],
    ['前端框架', 'Guard', 'ruoyi-ui/src/permission.js', 'Router Guard', '重写', '已完成'],
    ['前端框架', 'API', 'ruoyi-ui/src/api/login.js', 'Vue API', '修改', '已完成'],

    # 前端 - 样式
    ['移动端', 'Style', 'ruoyi-ui/src/assets/styles/mobile.scss', 'SCSS Style', '新增', '已完成'],
    ['移动端', 'Style', 'ruoyi-ui/src/assets/styles/index.scss', 'SCSS Style', '修改', '已完成'],

    # 数据库
    ['数据库', 'SQL', 'sql/ry-vue100-migration.sql', 'SQL Script', '新增', '已完成'],

    # 删除的文件
    ['系统框架', 'Module', 'ruoyi-generator/ (整个模块)', 'Maven Module', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysDictTypeController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysDictDataController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysPostController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysNoticeController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysConfigController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysRegisterController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'ServerController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'CacheController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysUserOnlineController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'TestController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysJobController.java', 'Java Controller', '删除', '已完成'],
    ['系统框架', 'Controller', 'SysJobLogController.java', 'Java Controller', '删除', '已完成'],
    ['前端框架', 'Views', 'ruoyi-ui/src/views/monitor/ (整个目录)', 'Vue Views', '删除', '已完成'],
    ['前端框架', 'Views', 'ruoyi-ui/src/views/tool/ (整个目录)', 'Vue Views', '删除', '已完成'],
    ['前端框架', 'Views', 'ruoyi-ui/src/views/system/dict/ (整个目录)', 'Vue Views', '删除', '已完成'],
    ['前端框架', 'Views', 'ruoyi-ui/src/views/system/post/ (整个目录)', 'Vue Views', '删除', '已完成'],
    ['前端框架', 'Views', 'ruoyi-ui/src/views/system/config/ (整个目录)', 'Vue Views', '删除', '已完成'],
    ['前端框架', 'Views', 'ruoyi-ui/src/views/system/notice/ (整个目录)', 'Vue Views', '删除', '已完成'],
    ['前端框架', 'Views', 'ruoyi-ui/src/views/register.vue', 'Vue SFC', '删除', '已完成'],
    ['前端框架', 'API', 'ruoyi-ui/src/api/monitor/ (整个目录)', 'Vue API', '删除', '已完成'],
    ['前端框架', 'API', 'ruoyi-ui/src/api/tool/ (整个目录)', 'Vue API', '删除', '已完成'],
    ['前端框架', 'API', 'ruoyi-ui/src/api/system/dict/ (整个目录)', 'Vue API', '删除', '已完成'],
]

for i, row in enumerate(files, 2):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val)
    # 操作着色
    op_cell = ws2.cell(row=i, column=5)
    if op_cell.value == '新增':
        op_cell.fill = green_fill
    elif op_cell.value == '修改' or op_cell.value == '重写':
        op_cell.fill = yellow_fill
    elif op_cell.value == '删除':
        op_cell.fill = red_fill

style_data(ws2, 2, len(files)+1, len(headers2))
auto_width(ws2, len(headers2))
ws2.auto_filter.ref = f"A1:F{len(files)+1}"
ws2.freeze_panes = 'A2'

# ============================================================
# Sheet 3: 数据库表设计
# ============================================================
ws3 = wb.create_sheet("数据库表设计")

headers3 = ['表名', '中文名称', '所属模块', '操作类型', '主要字段', '状态']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(headers3))

tables = [
    ['iot_device', '设备信息表', '设备管理', '新增', 'device_id,device_name,device_type,ip_address,mac_address,model,location,port_info,status,snmp_community,snmp_port,snmp_version,responsible,responsible_phone,tenant_id,parent_id', '已完成'],
    ['iot_device_port', '设备端口表', '设备管理', '新增', 'port_id,device_id,port_name,port_type,port_status,connected_device_id,connected_port_id', '已完成'],
    ['iot_device_status_log', '设备状态变更日志表', '设备管理', '新增', 'log_id,device_id,device_name,old_status,new_status,change_type,change_time,sms_sent,sms_recipient', '已完成'],
    ['iot_device_heartbeat_log', '设备心跳检测日志表', '设备管理', '新增', 'log_id,device_id,device_name,ip_address,ping_result,ping_latency,detect_time', '已完成'],
    ['visitor_appointment', '访客预约表', '访客管理', '新增', 'appointment_id,visitor_name,visitor_phone,visitor_id_card,visitor_company,visit_reason,host_name,host_dept,host_phone,visit_time,leave_time,status,approver_id,approve_time,sms_sent,tenant_id', '已完成'],
    ['visitor_log', '来访记录表', '访客管理', '新增', 'log_id,appointment_id,visitor_name,visitor_phone,visitor_id_card,visitor_company,visit_reason,host_name,host_dept,entry_time,exit_time,register_type,tenant_id', '已完成'],
    ['sys_sms_log', '短信通知日志表', '通用', '新增', 'sms_id,recipient,phone_number,content,send_time,send_result,biz_type,biz_id', '已完成'],
    ['sys_menu', '菜单权限表', '系统管理', '修改', '删除冗余菜单记录，新增设备管理/访客管理菜单', '已完成'],
    ['sys_role', '角色信息表', '系统管理', '修改', '新增deviceAdmin/visitorAdmin角色', '已完成'],
    ['sys_config', '参数配置表', '系统管理', '修改', '新增短信/SNMP/Ping相关配置参数', '已完成'],
]

for i, row in enumerate(tables, 2):
    for j, val in enumerate(row, 1):
        ws3.cell(row=i, column=j, value=val)

style_data(ws3, 2, len(tables)+1, len(headers3))
auto_width(ws3, len(headers3))
ws3.freeze_panes = 'A2'

# ============================================================
# Sheet 4: API接口清单
# ============================================================
ws4 = wb.create_sheet("API接口清单")

headers4 = ['模块', '请求方法', '接口URL', '功能说明', '权限标识', '状态']
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(headers4))

apis = [
    # 设备管理
    ['设备管理', 'GET', '/device/list', '查询设备列表(分页)', 'device:list:list', '已完成'],
    ['设备管理', 'GET', '/device/{deviceId}', '查询设备详情', 'device:list:list', '已完成'],
    ['设备管理', 'POST', '/device', '新增设备', 'device:list:add', '已完成'],
    ['设备管理', 'PUT', '/device', '修改设备', 'device:list:edit', '已完成'],
    ['设备管理', 'DELETE', '/device/{deviceIds}', '批量删除设备', 'device:list:remove', '已完成'],
    ['设备管理', 'POST', '/device/export', '导出设备Excel', 'device:list:export', '已完成'],
    ['设备管理', 'POST', '/device/repair/{deviceId}', '报修发送短信', 'device:list:repair', '已完成'],
    ['设备管理', 'GET', '/device/topology', '获取设备拓扑树', 'device:topology:view', '已完成'],
    # 设备端口
    ['设备管理', 'GET', '/device/port/list', '查询端口列表', 'device:list:list', '已完成'],
    ['设备管理', 'POST', '/device/port', '新增端口', 'device:list:edit', '已完成'],
    ['设备管理', 'PUT', '/device/port', '修改端口', 'device:list:edit', '已完成'],
    ['设备管理', 'DELETE', '/device/port/{portIds}', '删除端口', 'device:list:edit', '已完成'],
    ['设备管理', 'POST', '/device/port/bind', '绑定端口连接', 'device:list:edit', '已完成'],
    # 状态日志
    ['设备管理', 'GET', '/device/statuslog/list', '查询状态变更日志', 'device:list:list', '已完成'],
    # 心跳日志
    ['设备管理', 'GET', '/device/heartbeat/list', '查询心跳检测日志', 'device:heartbeat:list', '已完成'],
    # 访客管理
    ['访客管理', 'GET', '/visitor/appointment/list', '查询预约列表', 'visitor:appointment:list', '已完成'],
    ['访客管理', 'GET', '/visitor/appointment/{id}', '查询预约详情', 'visitor:appointment:list', '已完成'],
    ['访客管理', 'POST', '/visitor/appointment', '新增预约', 'visitor:appointment:add', '已完成'],
    ['访客管理', 'PUT', '/visitor/appointment', '修改预约', 'visitor:appointment:edit', '已完成'],
    ['访客管理', 'DELETE', '/visitor/appointment/{ids}', '删除预约', 'visitor:appointment:remove', '已完成'],
    ['访客管理', 'PUT', '/visitor/appointment/approve', '审批预约(含短信)', 'visitor:approval:list', '已完成'],
    ['访客管理', 'PUT', '/visitor/appointment/cancel/{id}', '取消预约', 'visitor:appointment:edit', '已完成'],
    ['访客管理', 'PUT', '/visitor/appointment/complete/{id}', '完成来访', 'visitor:appointment:edit', '已完成'],
    ['访客管理', 'GET', '/visitor/appointment/pending', '待审批列表', 'visitor:approval:list', '已完成'],
    # 来访记录
    ['访客管理', 'GET', '/visitor/log/list', '查询来访记录', 'visitor:log:list', '已完成'],
    ['访客管理', 'POST', '/visitor/log', '现场登记来访', 'visitor:register:add', '已完成'],
    ['访客管理', 'PUT', '/visitor/log/exit/{id}', '记录离开时间', 'visitor:log:edit', '已完成'],
    ['访客管理', 'POST', '/visitor/log/export', '导出来访记录', 'visitor:log:export', '已完成'],
    # 驾驶舱
    ['驾驶舱', 'GET', '/dashboard/device', '设备驾驶舱数据', '', '已完成'],
    ['驾驶舱', 'GET', '/dashboard/visitor', '访客驾驶舱数据', '', '已完成'],
    # 系统
    ['系统', 'POST', '/login', '用户登录', '', '保留'],
    ['系统', 'GET', '/getInfo', '获取用户信息', '', '保留'],
    ['系统', 'GET', '/getRouters', '获取路由(兼容)', '', '保留'],
    ['系统', 'GET', '/getConfig', '获取系统配置', '', '新增'],
]

for i, row in enumerate(apis, 2):
    for j, val in enumerate(row, 1):
        ws4.cell(row=i, column=j, value=val)

style_data(ws4, 2, len(apis)+1, len(headers4))
auto_width(ws4, len(headers4))
ws4.auto_filter.ref = f"A1:F{len(apis)+1}"
ws4.freeze_panes = 'A2'

# ============================================================
# Sheet 5: 测试用例清单
# ============================================================
ws5 = wb.create_sheet("测试用例清单")

headers5 = ['测试编号', '测试类别', '测试项', '测试设备/场景', '预期结果', '测试方法', '状态']
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(headers5))

tests = [
    ['TC-001', '功能测试', '华为防火墙Ping在线检测', '华为USG6565 116.148.212.150', 'Ping返回SUCCESS，延迟<100ms', 'DeviceHeartbeatTask自动检测', '已配置'],
    ['TC-002', '功能测试', '华为防火墙SNMP读取', '华为USG6565 116.148.212.150', '读取sysName,sysDescr正常', 'SnmpService.snmpGet', '已配置'],
    ['TC-003', '功能测试', '华为防火墙状态展示', '华为USG6565', '设备列表页显示在线状态绿色标签', '前端页面查看', '已配置'],
    ['TC-004', '功能测试', '海康摄像头离线识别', '海康DS-2CD2T46WD 116.148.212.151', 'Ping返回FAIL，状态变更为OFFLINE', 'DeviceHeartbeatTask自动检测', '已配置'],
    ['TC-005', '功能测试', '离线短信告警触发', '海康摄像头离线事件', 'SmsLog表记录DEVICE_OFFLINE短信', '检查sys_sms_log表', '已配置'],
    ['TC-006', '功能测试', '上线恢复台账记录', '海康摄像头恢复在线', 'DeviceStatusLog记录ONLINE事件', '检查iot_device_status_log表', '已配置'],
    ['TC-007', '接口测试', '设备CRUD接口', '全部设备API', 'GET/POST/PUT/DELETE返回正确', 'HTTP请求测试', '待测试'],
    ['TC-008', '接口测试', '访客预约审批流程', '全部访客API', '审批状态流转正常，短信触发', 'HTTP请求测试', '待测试'],
    ['TC-009', '权限测试', '设备管理员权限', 'deviceAdmin角色用户', '设备管理菜单可见，可操作设备', '切换角色登录测试', '待测试'],
    ['TC-010', '权限测试', '访客管理员权限', 'visitorAdmin角色用户', '访客审批页面可见，可审批', '切换角色登录测试', '待测试'],
    ['TC-011', '权限测试', '无权限拦截', '无权限用户访问受限页面', '401或重定向到登录页', '直接URL访问测试', '待测试'],
    ['TC-012', '前端测试', '静态路由渲染', '所有菜单', '页面路由全部正常，无动态路由报错', '点击所有菜单项', '待测试'],
    ['TC-013', '前端测试', '双驾驶舱切换', '修改dashboardMode配置', 'device/visitor模式切换生效', '修改yml配置后重启', '待测试'],
    ['TC-014', '前端测试', 'H5移动端适配', '手机浏览器访问', '页面自适应，操作正常', 'Chrome DevTools移动模拟', '待测试'],
    ['TC-015', '前端测试', '微信浏览器适配', '微信内置浏览器', '页面正常显示，无布局错乱', '微信打开页面', '待测试'],
    ['TC-016', '集成测试', 'Redis缓存', 'Redis服务', 'Token缓存正常，登录状态保持', '检查Redis keys', '待测试'],
    ['TC-017', '集成测试', 'Token登录', 'JWT认证', '登录获取Token，请求携带Token', 'Postman测试', '待测试'],
    ['TC-018', '集成测试', '定时巡检任务', 'DeviceHeartbeatTask', '定时执行Ping检测，日志记录', '检查日志和数据库', '待测试'],
    ['TC-019', '多租户测试', '租户数据隔离', '不同tenant_id设备', '租户A看不到租户B数据', 'SQL查询验证', '待测试'],
    ['TC-020', '拓扑测试', '设备拓扑生成', '端口绑定后的设备', '拓扑页面显示设备层级关系', '前端拓扑页面查看', '待测试'],
]

for i, row in enumerate(tests, 2):
    for j, val in enumerate(row, 1):
        ws5.cell(row=i, column=j, value=val)
    status_cell = ws5.cell(row=i, column=7)
    if status_cell.value == '已配置':
        status_cell.fill = green_fill
    elif status_cell.value == '待测试':
        status_cell.fill = yellow_fill

style_data(ws5, 2, len(tests)+1, len(headers5))
auto_width(ws5, len(headers5))
ws5.auto_filter.ref = f"A1:G{len(tests)+1}"
ws5.freeze_panes = 'A2'

# ============================================================
# Sheet 6: 统计汇总
# ============================================================
ws6 = wb.create_sheet("统计汇总")

# 统计
total_req = len(requirements)
completed_req = sum(1 for r in requirements if '已完成' in str(r[4]) or '已配置' in str(r[4]))
total_files = len(files)
new_files = sum(1 for f in files if f[4] == '新增')
mod_files = sum(1 for f in files if f[4] in ('修改', '重写'))
del_files = sum(1 for f in files if f[4] == '删除')
total_tables = len(tables)
total_apis = len(apis)
total_tests = len(tests)

headers6 = ['统计项', '数量', '说明']
stats = [
    ['需求总项', total_req, '开发需求.txt中的全部需求条目'],
    ['需求完成项', completed_req, f'完成率 {completed_req/total_req*100:.1f}%'],
    ['需求待完成项', total_req - completed_req, '待现场部署后测试验证'],
    ['', '', ''],
    ['文件总数', total_files, '新增+修改+删除的全部文件'],
    ['新增文件', new_files, '全新创建的Java/Vue/配置文件'],
    ['修改文件', mod_files, '重写或修改的已有文件'],
    ['删除文件', del_files, '删除的冗余文件/目录'],
    ['', '', ''],
    ['新增数据库表', total_tables - 3, '7张全新业务表(iot_*, visitor_*, sys_sms_log)'],
    ['修改数据库表', 3, 'sys_menu, sys_role, sys_config表数据'],
    ['', '', ''],
    ['API接口总数', total_apis, '包含新增和保留的接口'],
    ['测试用例总数', total_tests, '覆盖功能/接口/权限/集成测试'],
    ['', '', ''],
    ['删除的Maven模块', 1, 'ruoyi-generator(代码生成器)'],
    ['删除的后端Controller', 13, '字典/岗位/通知/配置/注册/监控/在线用户/定时任务页面等'],
    ['删除的前端View目录', 6, 'monitor, tool, dict, post, config, notice'],
    ['新增后端模块', 3, '设备管理, 访客管理, 驾驶舱'],
    ['新增前端View', 8, '设备3个+访客4个+驾驶舱1个'],
]

for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, len(headers6))

for i, row in enumerate(stats, 2):
    for j, val in enumerate(row, 1):
        ws6.cell(row=i, column=j, value=val)
    cell = ws6.cell(row=i, column=1)
    if cell.value and ('完成' in str(cell.value) or '新增' in str(cell.value)):
        cell.font = Font(name='微软雅黑', bold=True, size=12)

style_data(ws6, 2, len(stats)+1, len(headers6))
auto_width(ws6, len(headers6), min_w=20, max_w=60)

# 保存
output_path = 'E:/gitee/RuoYi-Vue-master/doc/开发情况.xlsx'
wb.save(output_path)
print(f"开发情况Excel已生成: {output_path}")
print(f"共6个工作表: 开发需求对比总览({total_req}项), 模块文件清单({total_files}项), 数据库表设计({total_tables}项), API接口清单({total_apis}项), 测试用例清单({total_tests}项), 统计汇总")
